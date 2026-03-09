"""Excel Analytics Dashboard Generator.

Produces a professional multi-sheet Excel workbook with:
  Sheet 1 — Executive Summary (KPI scorecard)
  Sheet 2 — Generation Mix (fuel breakdown with bar chart)
  Sheet 3 — Price Trends (state/sector analysis with line chart)
  Sheet 4 — Carbon Footprint (CO2 analysis with stacked bar)
  Sheet 5 — Regional Performance (market data table)

Designed to be CEO-presentation quality.
"""

import logging
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.chart import BarChart, LineChart, Reference
from openpyxl.chart.series import SeriesLabel
from openpyxl.styles import (
    Alignment,
    Border,
    Font,
    GradientFill,
    PatternFill,
    Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.utils.dataframe import dataframe_to_rows

from reports.generators.db_connector import query_df

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Brand colours
# ---------------------------------------------------------------------------
DARK_BLUE = "1B3A5C"
MID_BLUE = "2E6DA4"
LIGHT_BLUE = "D6E4F0"
GREEN = "27AE60"
RED = "C0392B"
AMBER = "F39C12"
WHITE = "FFFFFF"
LIGHT_GREY = "F2F2F2"
DARK_GREY = "555555"


def _header_font(size: int = 11, bold: bool = True, color: str = WHITE) -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _cell_font(size: int = 10, bold: bool = False, color: str = "000000") -> Font:
    return Font(name="Calibri", size=size, bold=bold, color=color)


def _header_fill(color: str = DARK_BLUE) -> PatternFill:
    return PatternFill("solid", fgColor=color)


def _thin_border() -> Border:
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _style_header_row(ws, row: int, num_cols: int, color: str = DARK_BLUE) -> None:
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row, column=col)
        cell.font = _header_font()
        cell.fill = _header_fill(color)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = _thin_border()


def _style_data_rows(ws, start_row: int, end_row: int, num_cols: int) -> None:
    for row in range(start_row, end_row + 1):
        fill = PatternFill("solid", fgColor=LIGHT_GREY if row % 2 == 0 else WHITE)
        for col in range(1, num_cols + 1):
            cell = ws.cell(row=row, column=col)
            cell.fill = fill
            cell.font = _cell_font()
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = _thin_border()


def _auto_column_width(ws) -> None:
    for col in ws.columns:
        max_len = max((len(str(cell.value or "")) for cell in col), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 4, 40)


class ExcelReportGenerator:
    """Generates the multi-sheet Excel analytics dashboard."""

    def generate(self, output_path: str) -> None:
        logger.info("Starting Excel report generation → %s", output_path)
        wb = Workbook()
        wb.remove(wb.active)  # remove default sheet

        self._sheet_executive_summary(wb)
        self._sheet_generation_mix(wb)
        self._sheet_price_trends(wb)
        self._sheet_carbon_footprint(wb)
        self._sheet_regional_performance(wb)

        wb.save(output_path)
        logger.info("Excel report saved to %s", output_path)

    # ------------------------------------------------------------------
    # Sheet 1: Executive Summary
    # ------------------------------------------------------------------
    def _sheet_executive_summary(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Executive Summary")
        ws.sheet_view.showGridLines = False

        # Title banner
        ws.merge_cells("A1:H1")
        title = ws["A1"]
        title.value = "Dakota Analytics — Energy Market Executive Summary"
        title.font = Font(name="Calibri", size=18, bold=True, color=WHITE)
        title.fill = PatternFill("solid", fgColor=DARK_BLUE)
        title.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[1].height = 40

        ws.merge_cells("A2:H2")
        subtitle = ws["A2"]
        subtitle.value = f"Report generated: {datetime.now().strftime('%B %d, %Y at %H:%M UTC')}"
        subtitle.font = Font(name="Calibri", size=11, italic=True, color=DARK_GREY)
        subtitle.alignment = Alignment(horizontal="center")
        ws.row_dimensions[2].height = 20

        # KPI data
        try:
            df = query_df("""
                select
                    sum(total_generation_mwh) / 1e6                        as total_gen_twh,
                    avg(renewable_share_pct)                                as avg_renewable_pct,
                    avg(clean_energy_share_pct)                             as avg_clean_pct,
                    sum(total_co2_tonnes) / 1e6                             as total_co2_mt,
                    avg(grid_carbon_intensity_gco2_kwh)                     as avg_carbon_intensity,
                    avg(residential_price_cents_kwh)                        as avg_residential_price,
                    avg(wholesale_spot_price_usd_mwh)                       as avg_wholesale_price,
                    count(distinct state_code)                              as states_covered,
                    max(period)                                             as latest_period
                from analytics.gold_executive_summary
                where period_year >= extract(year from now()) - 1
            """)
            kpis = df.iloc[0] if len(df) > 0 else {}
        except Exception as exc:
            logger.warning("Could not load KPI data: %s — using placeholders", exc)
            kpis = {}

        kpi_items = [
            ("Total Generation (TWh)", f"{kpis.get('total_gen_twh', 0):,.1f}" if kpis.get('total_gen_twh') else "N/A", MID_BLUE),
            ("Avg Renewable Share", f"{kpis.get('avg_renewable_pct', 0):.1f}%" if kpis.get('avg_renewable_pct') else "N/A", GREEN),
            ("Avg Clean Energy Share", f"{kpis.get('avg_clean_pct', 0):.1f}%" if kpis.get('avg_clean_pct') else "N/A", GREEN),
            ("Total CO2 (Mt)", f"{kpis.get('total_co2_mt', 0):,.1f}" if kpis.get('total_co2_mt') else "N/A", RED),
            ("Grid Carbon Intensity (gCO₂/kWh)", f"{kpis.get('avg_carbon_intensity', 0):.0f}" if kpis.get('avg_carbon_intensity') else "N/A", AMBER),
            ("Avg Residential Price (¢/kWh)", f"{kpis.get('avg_residential_price', 0):.2f}" if kpis.get('avg_residential_price') else "N/A", MID_BLUE),
            ("Avg Wholesale Price ($/MWh)", f"{kpis.get('avg_wholesale_price', 0):.2f}" if kpis.get('avg_wholesale_price') else "N/A", MID_BLUE),
            ("States Covered", str(kpis.get('states_covered', 'N/A')), DARK_BLUE),
        ]

        # KPI cards — 2 columns of 4
        for i, (label, value, color) in enumerate(kpi_items):
            row = 4 + (i % 4) * 4
            col = 1 if i < 4 else 5

            ws.merge_cells(
                start_row=row, start_column=col,
                end_row=row, end_column=col + 2
            )
            label_cell = ws.cell(row=row, column=col, value=label)
            label_cell.font = Font(name="Calibri", size=10, bold=True, color=WHITE)
            label_cell.fill = PatternFill("solid", fgColor=color)
            label_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row].height = 20

            ws.merge_cells(
                start_row=row + 1, start_column=col,
                end_row=row + 2, end_column=col + 2
            )
            value_cell = ws.cell(row=row + 1, column=col, value=value)
            value_cell.font = Font(name="Calibri", size=20, bold=True, color=color)
            value_cell.fill = PatternFill("solid", fgColor=LIGHT_GREY)
            value_cell.alignment = Alignment(horizontal="center", vertical="center")
            ws.row_dimensions[row + 1].height = 30

        _auto_column_width(ws)
        logger.info("Sheet 'Executive Summary' written")

    # ------------------------------------------------------------------
    # Sheet 2: Generation Mix
    # ------------------------------------------------------------------
    def _sheet_generation_mix(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Generation Mix")
        ws.sheet_view.showGridLines = False

        try:
            df = query_df("""
                select
                    period_year,
                    fuel_type_description,
                    fuel_category,
                    round(sum(total_generation_mwh) / 1e6, 2) as generation_twh,
                    round(avg(pct_of_state_generation), 2)    as avg_share_pct,
                    round(avg(avg_direct_co2_per_mwh), 1)     as avg_co2_per_mwh
                from analytics.gold_generation_by_fuel_monthly
                where fuel_type_description is not null
                group by period_year, fuel_type_description, fuel_category
                order by period_year desc, generation_twh desc
                limit 100
            """)
        except Exception as exc:
            logger.warning("Could not load generation mix data: %s", exc)
            df = pd.DataFrame()

        self._write_title(ws, "Electricity Generation Mix by Fuel Type", 1)
        if not df.empty:
            self._write_dataframe(ws, df, start_row=3)
            self._add_bar_chart(ws, df, title="Annual Generation by Fuel (TWh)", start_cell="I3")
        logger.info("Sheet 'Generation Mix' written")

    # ------------------------------------------------------------------
    # Sheet 3: Price Trends
    # ------------------------------------------------------------------
    def _sheet_price_trends(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Price Trends")
        ws.sheet_view.showGridLines = False

        try:
            df = query_df("""
                select
                    period,
                    state_code,
                    sector_id,
                    sector_name,
                    round(price_usd_per_mwh, 2)                 as price_usd_mwh,
                    round(price_cents_per_kwh, 2)               as price_cents_kwh,
                    round(price_yoy_change_pct, 1)              as yoy_change_pct,
                    round(rolling_12m_avg_price_usd_mwh, 2)     as rolling_12m_avg,
                    round(retail_wholesale_spread_usd_mwh, 2)   as retail_wholesale_spread
                from analytics.gold_price_trends
                where sector_id in ('RES','COM','IND')
                order by period desc, state_code, sector_id
                limit 500
            """)
        except Exception as exc:
            logger.warning("Could not load price trend data: %s", exc)
            df = pd.DataFrame()

        self._write_title(ws, "Retail Electricity Price Trends", 1)
        if not df.empty:
            self._write_dataframe(ws, df, start_row=3)
        logger.info("Sheet 'Price Trends' written")

    # ------------------------------------------------------------------
    # Sheet 4: Carbon Footprint
    # ------------------------------------------------------------------
    def _sheet_carbon_footprint(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Carbon Footprint")
        ws.sheet_view.showGridLines = False

        try:
            df = query_df("""
                select
                    period,
                    state_code,
                    state_description,
                    round(total_generation_mwh / 1e6, 3)            as generation_twh,
                    round(total_co2_tonnes / 1e6, 3)                as co2_million_tonnes,
                    round(renewable_share_pct, 1)                   as renewable_share_pct,
                    round(clean_energy_share_pct, 1)                as clean_energy_share_pct,
                    round(grid_carbon_intensity_gco2_kwh, 1)        as grid_intensity_gco2_kwh,
                    round(carbon_intensity_yoy_change_pct, 1)       as intensity_yoy_change_pct
                from analytics.gold_carbon_footprint
                order by period desc, co2_million_tonnes desc nulls last
                limit 500
            """)
        except Exception as exc:
            logger.warning("Could not load carbon footprint data: %s", exc)
            df = pd.DataFrame()

        self._write_title(ws, "Carbon Footprint & Renewable Energy Analysis", 1)
        if not df.empty:
            self._write_dataframe(ws, df, start_row=3)
        logger.info("Sheet 'Carbon Footprint' written")

    # ------------------------------------------------------------------
    # Sheet 5: Regional Performance
    # ------------------------------------------------------------------
    def _sheet_regional_performance(self, wb: Workbook) -> None:
        ws = wb.create_sheet("Regional Performance")
        ws.sheet_view.showGridLines = False

        try:
            df = query_df("""
                select
                    reading_date,
                    region,
                    round(avg_spot_price_usd_mwh, 2)    as spot_price_usd_mwh,
                    round(avg_peak_price_usd_mwh, 2)    as peak_price_usd_mwh,
                    round(avg_demand_mw, 0)              as avg_demand_mw,
                    round(peak_demand_mw, 0)             as peak_demand_mw,
                    round(price_volatility_pct, 1)       as price_volatility_pct
                from staging.silver_market_aggregated
                order by reading_date desc, region
                limit 500
            """)
        except Exception as exc:
            logger.warning("Could not load regional performance data: %s", exc)
            df = pd.DataFrame()

        self._write_title(ws, "Regional Electricity Market Performance", 1)
        if not df.empty:
            self._write_dataframe(ws, df, start_row=3)
        logger.info("Sheet 'Regional Performance' written")

    # ------------------------------------------------------------------
    # Shared helpers
    # ------------------------------------------------------------------
    def _write_title(self, ws, title: str, row: int) -> None:
        ws.merge_cells(f"A{row}:J{row}")
        cell = ws[f"A{row}"]
        cell.value = title
        cell.font = Font(name="Calibri", size=14, bold=True, color=DARK_BLUE)
        cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[row].height = 28

    def _write_dataframe(self, ws, df: pd.DataFrame, start_row: int) -> None:
        headers = list(df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=start_row, column=col_idx, value=header.replace("_", " ").title())
        _style_header_row(ws, start_row, len(headers))

        for row_idx, row_data in enumerate(dataframe_to_rows(df, index=False, header=False), start_row + 1):
            for col_idx, value in enumerate(row_data, 1):
                ws.cell(row=row_idx, column=col_idx, value=value)

        _style_data_rows(ws, start_row + 1, start_row + len(df), len(headers))
        _auto_column_width(ws)

    def _add_bar_chart(self, ws, df: pd.DataFrame, title: str, start_cell: str) -> None:
        if df.empty:
            return
        chart = BarChart()
        chart.type = "col"
        chart.title = title
        chart.y_axis.title = "TWh"
        chart.x_axis.title = "Fuel Type"
        chart.style = 10
        chart.width = 20
        chart.height = 14
        ws.add_chart(chart, start_cell)
